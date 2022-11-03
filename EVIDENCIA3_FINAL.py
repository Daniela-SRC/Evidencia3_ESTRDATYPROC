import sys
import sqlite3
from sqlite3 import Error
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import csv
import os
from os import path
import datetime
fecha_actual = datetime.date.today()
dia_actual = fecha_actual.day
mes_actual = fecha_actual.month
año_actual = fecha_actual.year
tupla_actual = (dia_actual, mes_actual, año_actual)
opcion = 0
clave_cliente = 0
clave_sala = 0
clave_registro = 0
fechaExistente = False
row=2

def menu():
    opc = int(input("Menú Principal\n" +
                    "Seleccione la opcion que guste:\n"+
                    "1.- Reservar\n" +
                    "2.- Reportes\n" +
                    "3.- Registrar un nuevo cliente\n" +
                    "4.- Registrar una sala\n" +
                    "5.- Finalizar\n"))
    return opc

def menu_reserva():
    opc1 = int(input("Seleccione la opcion que guste:\n"+
                    "1.- Registrar una reservación\n" +
                    "2.- Editar el nombre de un evento reservado\n" +
                    "3.- Consultar disponibilidad de salas\n" +
                    "4.- Eliminar una reservacion\n" +
                    "5.- Volver al menu principal\n"))
    return opc1

def menu_reporte():
    opc2 = int(input("Seleccione la opcion que guste:\n"+
                    "1.- Reporte de reservaciones para una fecha\n" +
                    "2.- Exportar reporte a Excel\n" +
                    "3.- Volver al menu principal\n"))
    return opc2


salas=[]
clientes=[]
eventos=[]
total_salas_turnos=[]
salas_turnos_ocupados=[]

try:
    with sqlite3.connect("bd_eventos.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS clientes (clave_c INTEGER PRIMARY KEY, nombre TEXT NOT NULL, apellidos TEXT NOT NULL);")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS salas (clave_s INTEGER PRIMARY KEY, nombre TEXT NOT NULL, cupo INTEGER NOT NULL);")
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS reservaciones (clave INTEGER PRIMARY KEY,nombre TEXT NOT NULL,turno TEXT NOT NULL,fecha_ev timestamp,cve_sala INTEGER NOT NULL,cve_cliente INTEGER NOT NULL,FOREIGN KEY(cve_sala) REFERENCES salas(clave_s),FOREIGN KEY(cve_cliente) REFERENCES clientes(clave_c));")
except Error as e:
    print (e)
except:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    
while opcion !=5:
    opcion = menu()
    if opcion == 1:
        opcion1 = menu_reserva()
        if opcion1 == 1:
            print("Registrar la reservación de una sala para un evento\n")
            if salas:
                clienteRegistrado = False            
                clave=int(input('Ingrese su ID: '))
                for elementoCliente in clientes:
                    for validacionID in range(len(elementoCliente)):        
                        if clave == elementoCliente[0]:            
                            clienteRegistrado = True
                            break
                        else:                
                            break
                if clienteRegistrado:
                    while True:
                        salaExistente = False
                        nombre_evento = input("Ingrese el nombre del evento: ") 
                        if nombre_evento != "": 
                            disponible = True 
                            cve_sala=int(input("Ingrese la clave de la sala del evento: ")) 
                            for revisionSala in salas:
                                for revisionClaveSala in range(len(revisionSala)):
                                    if cve_sala == revisionSala[0]:
                                        salaExistente = True                                    
                                        break                                
                            if salaExistente:                    
                                for Lista in eventos: 
                                    if disponible: 
                                        for claveIteracion in range(len(Lista)):                                                                                                                        
                                            if cve_sala == Lista[3]: 
                                                disponible = False
                                                break
                                    else:
                                        break                                
                                if disponible: 
                                    print("Continue con el registro")
                                    while True:
                                        try:
                                            horario_evento = int(input("Ingrese el numero (1,2 ó 3) del horario del evento que desee (1.-MATUTINO, 2.-VESPERTINO, 3.-NOCTURNO): "))
                                            
                                        
                                        except ValueError:
                                            print("Formato de dato incorrecto")
                                        else:
                                            if horario_evento == 1:
                                                turno_a_guardar="Matutino"
                                                
                                            elif horario_evento == 2:
                                                turno_a_guardar="Vespertino"
                                                
                                            elif horario_evento == 3:
                                                turno_a_guardar="Nocturno"
                                            
                                            if horario_evento > 0 and horario_evento < 4:
                                                print("Horario guardado")
                                                break
                                            else:
                                                print("Horario no valido")
                                                
                                    while True:
                                        fecha_reservada = input("Ingrese la fecha que desea reservar (dd/mm/aaaa): ")
                                        fecha_procesada = datetime.datetime.strptime(fecha_reservada,"%d/%m/%Y").date()
                                        dia_reservado = fecha_procesada.day
                                        mes_reservado = fecha_procesada.month
                                        año_reservado = fecha_procesada.year

                                        dia_valido = dia_reservado - dia_actual

                                        tupla_reservacion = (dia_reservado, mes_reservado, año_reservado)
                                        
                                        if dia_valido <= 1:
                                            print("Para reservar una fecha debe hacerlo con 2 dias de anticipación")
                                        else:
                                            if tupla_reservacion > tupla_actual:
                                                clave_registro += 1 
                                                print("Su reservación a sido éxitosa\n") 
                                                eventos.append((clave, clave_registro, nombre_evento, cve_sala, turno_a_guardar, tupla_reservacion))
                                                try:
                                                    with sqlite3.connect("bd_eventos.db") as conn:
                                                        mi_cursor = conn.cursor()
                                                        valores = {"nombre": nombre_evento, "turno": turno_a_guardar, "fecha_ev": fecha_procesada, "clave_s": cve_sala, "cliente": clave}
                                                        mi_cursor.execute("INSERT INTO reservaciones (nombre, turno, fecha_ev, cve_sala, cve_cliente) VALUES(:nombre, :turno, :fecha_ev, :clave_s, :cliente)", valores)
                                                except Error as e:
                                                    print (e)
                                                except:
                                                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                                finally:
                                                    conn.close()
                                                recorr=0
                                                for clave_sala, nombre_sala, cupo_sala in salas:                                    
                                                    if cve_sala ==salas[recorr][0]:
                                                        salita=nombre_sala
                                                        salas_turnos_ocupados.append((salita, turno_a_guardar)) 
                                                    recorr = recorr + 1
                                                    
                                                break
                                            else:
                                                print("Para reservar una fecha debe hacerlo con 2 dias de anticipación")
                                    break
                                else:
                                    print("ERROR! La sala ya ha sido registrada")                                
                            else:
                                print("ERROR! No existe esa sala")                            
                        else:
                             break                                
                else:
                    print("El cliente no está registrado")                                        
            else:
                print("ERROR! NO SE HA REGISTRADO ALGUNA SALA")
                
        if opcion1 == 2:
            print("Editar el nombre de un evento reservado\n")
            editar=int(input("Ingrese el ID de la reservacion que quiera modificar: "))
            nombre_ev=input("Ingrese el nuevo nombre del evento: ")
            try:
                with sqlite3.connect("bd_eventos.db") as conn:
                    mi_cursor = conn.cursor()
                    valores={"clave":editar, "nombre":nombre_ev}
                    mi_cursor.execute("UPDATE reservaciones SET nombre= :nombre WHERE clave = :clave", valores)
                    conn.commit()
                    print("Registro editado exitosamente\n")
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()
                
        if opcion1 == 3:
            print('Disponibilidad de salas\n')
            for clave_sala, nombre_sala, cupo_sala in salas:
                salita=nombre_sala
                total_salas_turnos.append((salita, "Matutino"))
                total_salas_turnos.append((salita, "Vespertino"))
                total_salas_turnos.append((salita, "Nocturno"))            
                conjunto_total_salas=set(total_salas_turnos)
                conjunto_salas_ocupadas=set(salas_turnos_ocupados)                     
                conjunto_salas_disponibles=conjunto_total_salas - conjunto_salas_ocupadas            
                print(conjunto_salas_disponibles)
            
        if opcion1 == 4:
            print("Eliminar una reservacion\n")
            eliminar=int(input("Ingrese el ID de la reservacion que quiera eliminar: "))
            try:
                with sqlite3.connect("bd_eventos.db") as conn:
                    mi_cursor = conn.cursor()
                    valores={"clave":eliminar}
                    mi_cursor.execute("DELETE FROM reservaciones WHERE  clave = :clave", valores)
                    conn.commit()
                    print("Registro eliminado\n")
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            finally:
                conn.close()
            
        if opcion1 == 5:
            print("-"*40)
        
    if opcion == 2:
        opcion2 = menu_reporte()
        if opcion2 == 1:
            print("Consulta de reservaciones\n")
            fechaExistente=False
            fecha_consulta = input("Ingrese la fecha que desea consultar (dd/mm/aaaa): ")
            fecha_consulta = datetime.datetime.strptime(fecha_consulta,"%d/%m/%Y").date()
            dia_consulta = fecha_consulta.day
            mes_consulta = fecha_consulta.month
            año_consulta = fecha_consulta.year            
            tupla_consulta = (dia_consulta, mes_consulta, año_consulta)
            print("--------------------------------------------------------------------")
            print(f"**\t\tREPORTE DE RESERVACIONES PARA EL DIA {fecha_consulta}\t\t**")
            print("--------------------------------------------------------------------")
            print("SALA\t CLIENTE\t\t EVENTO\t\t TURNO")
            print("--------------------------------------------------------------------")
            
            
            try:
                with sqlite3.connect("bd_eventos.db") as conn:
                    mi_cursor = conn.cursor()
                    valores_consulta={"fecha_consulta":fecha_consulta}
                    mi_cursor.execute("SELECT cve_sala, cve_cliente, nombre, turno FROM reservaciones where fecha_ev = :fecha_consulta", valores_consulta)
                    registro = mi_cursor.fetchall()
                    if registro:
                        for clave, clave_c, nombre, turno in registro:
                            print(f"{clave}\t{clave_c}\t{nombre}\t{turno}")
                    else:
                        print(f"No se encontraron reservaciones para el dia: {fecha_consulta}")
                    
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
            print("----------------------------FIN DEL REPORTE----------------------------")
            fechaExistente = True
        
        if opcion2 == 2:
            print("Reporte en Excel\n")
            fechaExistente=False
            fecha_consulta = input("Ingrese la fecha que desea consultar (dd/mm/aaaa): ")
            fecha_consulta = datetime.datetime.strptime(fecha_consulta,"%d/%m/%Y").date()
            dia_consulta = fecha_consulta.day
            mes_consulta = fecha_consulta.month
            año_consulta = fecha_consulta.year            
            tupla_consulta = (dia_consulta, mes_consulta, año_consulta)      
              
            libro = Workbook()
            hoja = libro.active
            hoja["A1"].value = "REPORTE DE EVENTOS PARA EL DIA: "
            hoja["B1"].value = fecha_consulta
            hoja["A2"].value= "SALA"
            hoja["B2"].value = "NOMBRE CLIENTE"
            hoja["C2"].value = "APELLIDO CLIENTE"
            hoja["D2"].value = "EVENTO"
            hoja["E2"].value = "TURNO"
            
            try:
                with sqlite3.connect("bd_eventos.db") as conn:
                    mi_cursor = conn.cursor()
                    valores_consulta={"fecha_consulta":fecha_consulta}
                    mi_cursor.execute("SELECT cve_sala, cve_cliente, nombre, turno FROM reservaciones where fecha_ev = :fecha_consulta", valores_consulta)
                    registro = mi_cursor.fetchall()
                    if registro:
                        
                        for clave, clave_c, nombre, turno in registro:
                            row=row+1
                            print("Revise su bandeja de archivos")
                            hoja.cell(row=row+1, column=1).value=clave
                            hoja.cell(row=row+1, column=2).value=clave_c
                            hoja.cell(row=row+1, column=3).value=nombre
                            hoja.cell(row=row+1, column=4).value=turno


                    else:
                        print(f"No se encontraron reservaciones para el dia: {fecha_consulta}")
                    
            except Error as e:
                print (e)
            except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                                
            fechaExistente = True
            
            libro.save('Consulta_eventos_prueba.xlsx')
            
        if opcion2 == 3:
            print("-"*40)
            
                
    if opcion == 3:
        print("Registrar un nuevo cliente\n")
        while True:
            nombre_cliente=input("Ingrese el nombre del cliente: ")
            if nombre_cliente == "":
                print("El nombre del cliente no puede omitirse\n")
            else:
                apellidos=input("Ingrese los apellidos del cliente: ")
                clave_cliente += 1
                print("Cliente agregado.\n")
                clientes.append((clave_cliente, nombre_cliente, apellidos))
                try:
                    with sqlite3.connect("bd_eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"nombre":nombre_cliente, "apellidos":apellidos}
                        mi_cursor.execute("INSERT INTO clientes (nombre, apellidos) VALUES(:nombre,:apellidos)", valores)
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                break
        
    if opcion == 4:
        print("Registrar una sala\n")
        while True:
            nombre_sala = input("Ingrese el nombre de la sala: ")
            if nombre_sala == "":
                print("El nombre de la sala no debe omitirse\n")
            else:
                cupo_sala = int(input("Ingrese el cupo de la sala: "))
                if cupo_sala <= 0:
                    print("El cupo de la sala debe ser un numero mayor a 0\n")
                else:
                    clave_sala += 1
                    salas.append((clave_sala, nombre_sala, cupo_sala))
                try:
                    with sqlite3.connect("bd_eventos.db") as conn:
                        mi_cursor = conn.cursor()
                        valores = {"nombre":nombre_sala, "cupo":cupo_sala}
                        mi_cursor.execute("INSERT INTO salas (nombre, cupo) VALUES(:nombre,:cupo)", valores)
                except Error as e:
                    print (e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    conn.close()
                    print("Sala agregada.\n")
                    break

    if opcion == 5:
        print("Usted a salido con éxito\n")
        break
    
    
    
    
    
    
    
 